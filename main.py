import openpyxl
import csv

# Open the CSV file for reading
with open('data.csv') as file:
    reader = csv.DictReader(file)
    lst_reader = list(reader)

    # Create a new Excel workbook and the active sheet
    book = openpyxl.Workbook()
    sheet = book.active

    # Populate column headers
    sheet['A2'] = 'Id'
    sheet.cell(row=3, column=1).value = 'Name'
    sheet[4][0].value = 'Phone'

    # Create column labels using a list iterator
    column_labels = [f"Person {i}" for i in range(1, len(lst_reader) + 1)]
    for col, label in zip(range(2, len(lst_reader) + 2), column_labels):
        sheet.cell(row=1, column=col).value = label

    # Populate data from the CSV file into the Excel sheet
    for idx, row in enumerate(lst_reader, start=2):
        sheet.cell(row=2, column=idx).value = row['Id']
        sheet.cell(row=3, column=idx).value = row['Name']
        sheet.cell(row=4, column=idx).value = row['Phone']

    # Save the Excel workbook as data.xlsx
    book.save('data.xlsx')
    # Close file
    book.close()
